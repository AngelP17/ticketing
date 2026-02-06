import os
import json
import hashlib
import secrets
from functools import wraps
from flask import Flask, render_template, jsonify, request, session, redirect, url_for, send_file
from datetime import datetime, time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import psycopg2
from psycopg2.extras import RealDictCursor

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'ticketing-dashboard-secret-key-2025')

# Database URL from environment (Render provides this)
DATABASE_URL = os.environ.get('DATABASE_URL')

# Check multiple locations for users file (local dev, Render secrets, Docker)
USERS_FILE_LOCATIONS = [
    '.env/users.json',              # Local development
    '/etc/secrets/users_data.json', # Render secret files
    'users.json',                   # Working directory fallback
]

def get_users_file():
    """Find the users file from multiple possible locations."""
    for path in USERS_FILE_LOCATIONS:
        if os.path.exists(path):
            return path
    return USERS_FILE_LOCATIONS[0]  # Default to first option

USERS_FILE = get_users_file()

# --- DATABASE FUNCTIONS ---
def clean_database_url(url):
    """Clean database URL by removing unsupported parameters for psycopg2."""
    if not url:
        return url

    from urllib.parse import urlparse, parse_qs, urlencode, urlunparse

    try:
        parsed = urlparse(url)

        # Parse query parameters
        params = parse_qs(parsed.query)

        # Remove unsupported parameters (like channel_binding from Neon)
        unsupported = ['channel_binding', 'options']
        for param in unsupported:
            params.pop(param, None)

        # Flatten params (parse_qs returns lists)
        clean_params = {k: v[0] if len(v) == 1 else v for k, v in params.items()}

        # Reconstruct URL
        new_query = urlencode(clean_params)
        clean_url = urlunparse((
            parsed.scheme,
            parsed.netloc,
            parsed.path,
            parsed.params,
            new_query,
            parsed.fragment
        ))

        print(f"[DB] Cleaned URL parameters: removed {[p for p in unsupported if p in parse_qs(parsed.query)]}")
        return clean_url
    except Exception as e:
        print(f"[DB] Warning: Could not parse URL, using as-is: {e}")
        return url

def get_db_connection():
    """Get a database connection."""
    if not DATABASE_URL:
        print("[DB] ERROR: DATABASE_URL not set!")
        return None
    try:
        clean_url = clean_database_url(DATABASE_URL)
        conn = psycopg2.connect(clean_url, cursor_factory=RealDictCursor)
        return conn
    except Exception as e:
        print(f"[DB] ERROR: Failed to connect to database: {e}")
        return None

def init_database():
    """Initialize the database tables."""
    conn = get_db_connection()
    if not conn:
        print("[DB] Cannot initialize database - no connection")
        return False

    try:
        cur = conn.cursor()
        cur.execute('''
            CREATE TABLE IF NOT EXISTS tickets (
                id SERIAL PRIMARY KEY,
                ticket_id VARCHAR(20) UNIQUE NOT NULL,
                title TEXT NOT NULL,
                status VARCHAR(50) DEFAULT 'Open',
                priority VARCHAR(50) DEFAULT 'Low',
                request_type VARCHAR(100),
                staff_assigned VARCHAR(100),
                requester VARCHAR(100),
                date_opened DATE DEFAULT CURRENT_DATE,
                description TEXT,
                resolution_notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Categories table (managed request types)
        cur.execute('''
            CREATE TABLE IF NOT EXISTS categories (
                id SERIAL PRIMARY KEY,
                name VARCHAR(100) UNIQUE NOT NULL,
                color VARCHAR(7) DEFAULT '#6366f1',
                icon VARCHAR(50) DEFAULT 'fa-tag',
                is_custom BOOLEAN DEFAULT FALSE,
                is_active BOOLEAN DEFAULT TRUE,
                sort_order INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Labels table (colored tags like Jira)
        cur.execute('''
            CREATE TABLE IF NOT EXISTS labels (
                id SERIAL PRIMARY KEY,
                name VARCHAR(50) UNIQUE NOT NULL,
                color VARCHAR(7) DEFAULT '#3b82f6',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Ticket-Labels junction table
        cur.execute('''
            CREATE TABLE IF NOT EXISTS ticket_labels (
                ticket_id VARCHAR(20) REFERENCES tickets(ticket_id) ON DELETE CASCADE,
                label_id INTEGER REFERENCES labels(id) ON DELETE CASCADE,
                PRIMARY KEY (ticket_id, label_id)
            )
        ''')

        # Ticket attachments table (images)
        cur.execute('''
            CREATE TABLE IF NOT EXISTS ticket_attachments (
                id SERIAL PRIMARY KEY,
                ticket_id VARCHAR(20) REFERENCES tickets(ticket_id) ON DELETE CASCADE,
                filename VARCHAR(255) NOT NULL,
                original_name VARCHAR(255) NOT NULL,
                mime_type VARCHAR(100) NOT NULL,
                file_data BYTEA NOT NULL,
                file_size INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Add category_id column to tickets if not exists
        cur.execute('''
            ALTER TABLE tickets ADD COLUMN IF NOT EXISTS category_id INTEGER REFERENCES categories(id)
        ''')

        conn.commit()
        print("[DB] Database tables initialized successfully")

        # Seed default categories
        default_categories = [
            ('Account Changes', '#3b82f6', 'fa-user-cog'),
            ('Account Setup', '#10b981', 'fa-user-plus'),
            ('Company Website Issues', '#f59e0b', 'fa-globe'),
            ('Documentation Request', '#8b5cf6', 'fa-file-alt'),
            ('Email Issues', '#ef4444', 'fa-envelope'),
            ('Feedback Request', '#ec4899', 'fa-comment'),
            ('Hardware Acquisition', '#6366f1', 'fa-shopping-cart'),
            ('Hardware Setup', '#818cf8', 'fa-desktop'),
            ('Network Problems', '#14b8a6', 'fa-wifi'),
            ('Permission Changes', '#f97316', 'fa-shield-alt'),
            ('Server Issues', '#dc2626', 'fa-server'),
            ('Software Acquisition', '#0ea5e9', 'fa-download'),
            ('Software Issues', '#38bdf8', 'fa-bug'),
            ('Software Setup', '#7dd3fc', 'fa-cog'),
            ('Spam Removal', '#84cc16', 'fa-trash'),
            ('System Access', '#a855f7', 'fa-lock'),
            ('Password Reset', '#06b6d4', 'fa-key'),
            ('Printer Issue', '#fb923c', 'fa-print'),
            ('VPN/Remote Access', '#4f46e5', 'fa-network-wired'),
            ('Security Incident', '#e11d48', 'fa-exclamation-triangle'),
            ('Database Support', '#a78bfa', 'fa-database'),
            ('Server Maintenance', '#d946ef', 'fa-tools'),
            ('Backup/Recovery', '#f472b6', 'fa-cloud-upload-alt'),
            ('Phone/VoIP', '#fb7185', 'fa-phone'),
            ('New Employee Setup', '#64748b', 'fa-user-plus'),
            ('Offboarding', '#475569', 'fa-user-minus'),
            ('General Inquiry', '#737373', 'fa-question-circle'),
            ('Change Request', '#a3a3a3', 'fa-exchange-alt'),
            ('Application Error', '#fb923c', 'fa-bug'),
        ]
        for name, color, icon in default_categories:
            cur.execute('''
                INSERT INTO categories (name, color, icon, is_custom)
                VALUES (%s, %s, %s, FALSE)
                ON CONFLICT (name) DO NOTHING
            ''', (name, color, icon))

        # Seed default labels
        default_labels = [
            ('BILLING', '#ef4444'),
            ('ACCOUNTS', '#3b82f6'),
            ('FORMS', '#22c55e'),
            ('FEEDBACK', '#f59e0b'),
            ('URGENT', '#dc2626'),
            ('FOLLOW-UP', '#8b5cf6'),
            ('ESCALATED', '#f97316'),
            ('SLA-BREACH', '#e11d48'),
        ]
        for name, color in default_labels:
            cur.execute('''
                INSERT INTO labels (name, color)
                VALUES (%s, %s)
                ON CONFLICT (name) DO NOTHING
            ''', (name, color))

        conn.commit()
        print("[DB] Default categories and labels seeded")

        # Migrate existing request_type values to categories
        cur.execute('''
            INSERT INTO categories (name, is_custom, is_active)
            SELECT DISTINCT request_type, FALSE, TRUE
            FROM tickets
            WHERE request_type IS NOT NULL
              AND request_type != ''
              AND request_type != 'nan'
            ON CONFLICT (name) DO NOTHING
        ''')

        # Link existing tickets to their categories
        cur.execute('''
            UPDATE tickets
            SET category_id = c.id
            FROM categories c
            WHERE tickets.request_type = c.name
              AND tickets.category_id IS NULL
        ''')

        conn.commit()
        print("[DB] Existing data migrated to categories")

        # Check if we need to migrate data from Excel
        cur.execute("SELECT COUNT(*) as count FROM tickets")
        count = cur.fetchone()['count']
        if count == 0:
            print("[DB] No tickets in database, attempting to migrate from Excel...")
            migrate_from_excel(conn)
        else:
            print(f"[DB] Found {count} existing tickets in database")

        cur.close()
        conn.close()
        return True
    except Exception as e:
        print(f"[DB] ERROR initializing database: {e}")
        import traceback
        traceback.print_exc()
        return False

def migrate_from_excel(conn):
    """Migrate existing data from Excel file to database."""
    try:
        from openpyxl import load_workbook
        EXCEL_FILE = 'tickets.xlsx'
        SHEET_NAME = 'IT Service Tickets'

        if not os.path.exists(EXCEL_FILE):
            print(f"[DB] Excel file {EXCEL_FILE} not found, skipping migration")
            return

        wb = load_workbook(EXCEL_FILE)
        ws = wb[SHEET_NAME]

        cur = conn.cursor()
        migrated = 0

        for row in range(2, 1000):
            title = ws[f'B{row}'].value
            if not title:
                break

            ticket_id = f"IT-2025{row-1:04d}"
            status = ws[f'C{row}'].value or 'Open'
            priority = ws[f'D{row}'].value or 'Low'
            request_type = ws[f'E{row}'].value or ''
            staff_assigned = ws[f'F{row}'].value or ''
            requester = ws[f'G{row}'].value or ''
            date_val = ws[f'H{row}'].value
            description = ws[f'J{row}'].value or ''
            resolution = ws[f'K{row}'].value or ''

            # Handle date conversion
            if date_val is None:
                date_opened = datetime.now().date()
            elif hasattr(date_val, 'date'):
                date_opened = date_val.date() if hasattr(date_val, 'date') else date_val
            elif hasattr(date_val, 'strftime'):
                date_opened = date_val
            else:
                try:
                    date_opened = datetime.strptime(str(date_val), '%Y-%m-%d').date()
                except:
                    date_opened = datetime.now().date()

            cur.execute('''
                INSERT INTO tickets (ticket_id, title, status, priority, request_type,
                                    staff_assigned, requester, date_opened, description, resolution_notes)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (ticket_id) DO NOTHING
            ''', (ticket_id, str(title).strip(), str(status).strip(), str(priority).strip(),
                  str(request_type).strip(), str(staff_assigned).strip(), str(requester).strip(),
                  date_opened, str(description).strip(), str(resolution).strip() if resolution else ''))
            migrated += 1

        conn.commit()
        cur.close()
        wb.close()
        print(f"[DB] Migrated {migrated} tickets from Excel to database")
    except Exception as e:
        print(f"[DB] ERROR during Excel migration: {e}")
        import traceback
        traceback.print_exc()

# --- USER MANAGEMENT ---
def load_users():
    """Load users from JSON file."""
    try:
        users_path = get_users_file()
        print(f"[AUTH] Loading users from: {users_path}")
        with open(users_path, 'r') as f:
            data = json.load(f)
            users = data.get('users', [])
            print(f"[AUTH] Loaded {len(users)} users successfully")
            return users
    except FileNotFoundError:
        print(f"[AUTH] ERROR: Users file not found at any location: {USERS_FILE_LOCATIONS}")
        return []
    except json.JSONDecodeError as e:
        print(f"[AUTH] ERROR: Invalid JSON in users file: {e}")
        return []
    except Exception as e:
        print(f"[AUTH] ERROR: Failed to load users: {e}")
        return []

def save_users(users):
    """Save users to JSON file."""
    with open(USERS_FILE, 'w') as f:
        json.dump({'users': users}, f, indent=2)

def hash_password(password):
    """Hash password using SHA256."""
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(password, password_hash):
    """Verify password against hash."""
    return hash_password(password) == password_hash

def get_user(username):
    """Get user by username."""
    users = load_users()
    for user in users:
        if user['username'] == username:
            return user
    return None

def login_required(f):
    """Decorator to require login."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user' not in session:
            return jsonify({'error': 'Not authenticated'}), 401
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    """Decorator to require admin role."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user' not in session:
            return jsonify({'error': 'Not authenticated'}), 401
        if session['user'].get('role') != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        return f(*args, **kwargs)
    return decorated

# --- DATABASE TICKET FUNCTIONS ---
def read_tickets_from_db():
    """Read all tickets from the database with category and label data."""
    conn = get_db_connection()
    if not conn:
        return []

    try:
        cur = conn.cursor()
        cur.execute('''
            SELECT t.ticket_id, t.title, t.status, t.priority, t.request_type,
                   t.staff_assigned, t.requester, t.date_opened, t.description,
                   t.resolution_notes, t.category_id,
                   c.name as category_name, c.color as category_color
            FROM tickets t
            LEFT JOIN categories c ON t.category_id = c.id
            ORDER BY t.date_opened DESC, t.id DESC
        ''')
        rows = cur.fetchall()

        # Fetch all ticket-label associations in one query
        cur.execute('''
            SELECT tl.ticket_id, l.id as label_id, l.name, l.color
            FROM ticket_labels tl
            JOIN labels l ON tl.label_id = l.id
        ''')
        label_rows = cur.fetchall()
        cur.close()
        conn.close()

        # Build label lookup
        ticket_labels = {}
        for lr in label_rows:
            tid = lr['ticket_id']
            if tid not in ticket_labels:
                ticket_labels[tid] = []
            ticket_labels[tid].append({'id': lr['label_id'], 'name': lr['name'], 'color': lr['color']})

        tickets = []
        for row in rows:
            date_opened = row['date_opened']
            if date_opened:
                date_str = date_opened.strftime('%Y-%m-%d') if hasattr(date_opened, 'strftime') else str(date_opened)
                if row['status'] in ['Closed', 'Resolved']:
                    days_open = -1
                else:
                    days_open = (datetime.now().date() - date_opened).days
                    if days_open < 0:
                        days_open = 0
            else:
                date_str = ''
                days_open = 0

            tickets.append({
                'ticket_id': row['ticket_id'],
                'title': row['title'] or '',
                'status': row['status'] or 'Open',
                'priority': row['priority'] or 'Low',
                'request_type': row.get('category_name') or row['request_type'] or '',
                'staff_assigned': row['staff_assigned'] or '',
                'requester': row['requester'] or '',
                'date_opened': date_str,
                'days_open': days_open,
                'description': row['description'] or '',
                'resolution_notes': row['resolution_notes'] or '',
                'category_id': row.get('category_id'),
                'category_color': row.get('category_color', '#6366f1'),
                'labels': ticket_labels.get(row['ticket_id'], [])
            })

        return tickets
    except Exception as e:
        print(f"[DB] ERROR reading tickets: {e}")
        import traceback
        traceback.print_exc()
        return []

def get_next_ticket_id():
    """Generate the next ticket ID."""
    conn = get_db_connection()
    if not conn:
        return 'IT-20250001'

    try:
        cur = conn.cursor()
        cur.execute("SELECT ticket_id FROM tickets ORDER BY id DESC LIMIT 1")
        row = cur.fetchone()
        cur.close()
        conn.close()

        if row:
            # Extract number from ticket_id (e.g., IT-20250001 -> 20250001)
            num = int(row['ticket_id'].replace('IT-', ''))
            return f"IT-{num + 1:08d}"
        return 'IT-20250001'
    except Exception as e:
        print(f"[DB] ERROR getting next ticket ID: {e}")
        return 'IT-20250001'

def create_ticket_in_db(ticket_data):
    """Create a new ticket in the database with category and label support."""
    conn = get_db_connection()
    if not conn:
        return None

    try:
        cur = conn.cursor()
        ticket_id = get_next_ticket_id()

        # Resolve category_id from request_type name if not provided
        category_id = ticket_data.get('category_id')
        request_type = ticket_data.get('request_type', '')
        if not category_id and request_type:
            cur.execute('SELECT id FROM categories WHERE name = %s', (request_type,))
            cat = cur.fetchone()
            if cat:
                category_id = cat['id']

        cur.execute('''
            INSERT INTO tickets (ticket_id, title, status, priority, request_type,
                                staff_assigned, requester, date_opened, description,
                                resolution_notes, category_id)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING ticket_id
        ''', (
            ticket_id,
            ticket_data.get('title', ''),
            ticket_data.get('status', 'Open'),
            ticket_data.get('priority', 'Low'),
            request_type,
            ticket_data.get('staff_assigned', ''),
            ticket_data.get('requester', ''),
            datetime.now().date(),
            ticket_data.get('description', ''),
            ticket_data.get('resolution_notes', ''),
            category_id
        ))

        result = cur.fetchone()

        # Handle labels
        label_ids = ticket_data.get('label_ids', [])
        for lid in label_ids:
            cur.execute('INSERT INTO ticket_labels (ticket_id, label_id) VALUES (%s, %s) ON CONFLICT DO NOTHING',
                        (ticket_id, int(lid)))

        conn.commit()
        cur.close()
        conn.close()

        print(f"[DB] Created ticket: {ticket_id}")
        return result['ticket_id']
    except Exception as e:
        print(f"[DB] ERROR creating ticket: {e}")
        import traceback
        traceback.print_exc()
        return None

def update_ticket_in_db(ticket_id, ticket_data):
    """Update a ticket in the database."""
    conn = get_db_connection()
    if not conn:
        return False

    try:
        cur = conn.cursor()

        # Build dynamic update query
        updates = []
        values = []

        field_mapping = {
            'title': 'title',
            'status': 'status',
            'priority': 'priority',
            'request_type': 'request_type',
            'staff_assigned': 'staff_assigned',
            'requester': 'requester',
            'description': 'description',
            'resolution_notes': 'resolution_notes'
        }

        for key, db_field in field_mapping.items():
            if key in ticket_data:
                updates.append(f"{db_field} = %s")
                values.append(ticket_data[key])

        if not updates:
            return True

        updates.append("updated_at = CURRENT_TIMESTAMP")
        values.append(ticket_id)

        query = f"UPDATE tickets SET {', '.join(updates)} WHERE ticket_id = %s"
        cur.execute(query, values)

        updated = cur.rowcount > 0
        conn.commit()
        cur.close()
        conn.close()

        print(f"[DB] Updated ticket: {ticket_id}, affected rows: {cur.rowcount}")
        return updated
    except Exception as e:
        print(f"[DB] ERROR updating ticket: {e}")
        import traceback
        traceback.print_exc()
        return False

def delete_ticket_from_db(ticket_id):
    """Delete a ticket from the database."""
    conn = get_db_connection()
    if not conn:
        return False

    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM tickets WHERE ticket_id = %s", (ticket_id,))
        deleted = cur.rowcount > 0
        conn.commit()
        cur.close()
        conn.close()

        print(f"[DB] Deleted ticket: {ticket_id}")
        return deleted
    except Exception as e:
        print(f"[DB] ERROR deleting ticket: {e}")
        return False

def calculate_stats(tickets):
    """Calculates KPIs from the ticket list."""
    total = len(tickets)
    open_count = sum(1 for t in tickets if t['status'] not in ['Closed', 'Resolved'])
    closed_count = sum(1 for t in tickets if t['status'] in ['Closed', 'Resolved'])
    critical = sum(1 for t in tickets if t['priority'] == 'Critical' and t['status'] not in ['Closed', 'Resolved'])

    statuses = {}
    for t in tickets:
        s = t['status']
        statuses[s] = statuses.get(s, 0) + 1

    priorities = {}
    for t in tickets:
        p = t['priority']
        priorities[p] = priorities.get(p, 0) + 1

    request_types = {}
    for t in tickets:
        rt = t['request_type']
        if rt and rt != 'nan':
            request_types[rt] = request_types.get(rt, 0) + 1

    staff_workload = {}
    for t in tickets:
        staff = t['staff_assigned']
        if staff and staff != 'nan':
            if staff not in staff_workload:
                staff_workload[staff] = {'assigned': 0, 'open': 0}
            staff_workload[staff]['assigned'] += 1
            if t['status'] not in ['Closed', 'Resolved']:
                staff_workload[staff]['open'] += 1

    return {
        'total': total,
        'open': open_count,
        'closed': closed_count,
        'critical': critical,
        'statuses': statuses,
        'priorities': priorities,
        'request_types': request_types,
        'staff_workload': staff_workload
    }

def get_dropdown_options():
    """Get unique values for dropdown fields from categories table and ticket data."""
    conn = get_db_connection()
    categories = []
    labels = []
    if conn:
        try:
            cur = conn.cursor()
            cur.execute('SELECT id, name, color, icon FROM categories WHERE is_active = TRUE ORDER BY sort_order, name')
            categories = [dict(c) for c in cur.fetchall()]
            cur.execute('SELECT id, name, color FROM labels ORDER BY name')
            labels = [dict(l) for l in cur.fetchall()]
            cur.close()
            conn.close()
        except Exception as e:
            print(f"[DB] ERROR in get_dropdown_options: {e}")

    tickets = read_tickets_from_db()
    return {
        'request_types': [c['name'] for c in categories],
        'categories': categories,
        'labels': labels,
        'staff': sorted(set(t['staff_assigned'] for t in tickets if t['staff_assigned'] and t['staff_assigned'] != 'nan')),
        'requesters': sorted(set(t['requester'] for t in tickets if t['requester'] and t['requester'] != 'nan'))
    }

# --- EXCEL EXPORT FUNCTION ---
def generate_excel_from_db():
    """Generate a professionally styled Excel file from database tickets."""
    tickets = read_tickets_from_db()

    wb = Workbook()

    # --- Sheet 1: Ticket Data ---
    ws = wb.active
    ws.title = "IT Service Tickets"

    headers = ['Ticket ID', 'Title', 'Status', 'Priority', 'Request Type',
               'Labels', 'Staff Assigned', 'Requester', 'Date Opened', 'Days Open',
               'Description', 'Resolution Notes']

    # Header styling
    header_fill = PatternFill(start_color='1a1a1a', end_color='1a1a1a', fill_type='solid')
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='404040'),
        right=Side(style='thin', color='404040'),
        top=Side(style='thin', color='404040'),
        bottom=Side(style='thin', color='404040')
    )

    # Write and style headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Priority cell fills
    priority_fills = {
        'Critical': PatternFill(start_color='fecaca', end_color='fecaca', fill_type='solid'),
        'High': PatternFill(start_color='fed7aa', end_color='fed7aa', fill_type='solid'),
        'Medium': PatternFill(start_color='fef3c7', end_color='fef3c7', fill_type='solid'),
        'Low': PatternFill(start_color='dcfce7', end_color='dcfce7', fill_type='solid'),
    }
    priority_fonts = {
        'Critical': Font(color='991B1B', bold=True),
        'High': Font(color='9A3412'),
        'Medium': Font(color='92400E'),
        'Low': Font(color='166534'),
    }

    # Status cell fills
    status_fills = {
        'Open': PatternFill(start_color='fecaca', end_color='fecaca', fill_type='solid'),
        'In Progress': PatternFill(start_color='dbeafe', end_color='dbeafe', fill_type='solid'),
        'Waiting for Info': PatternFill(start_color='fef3c7', end_color='fef3c7', fill_type='solid'),
        'Resolved': PatternFill(start_color='ccfbf1', end_color='ccfbf1', fill_type='solid'),
        'Closed': PatternFill(start_color='dcfce7', end_color='dcfce7', fill_type='solid'),
    }

    # Data rows
    for row_idx, t in enumerate(tickets, 2):
        labels_str = ', '.join(l['name'] for l in t.get('labels', []))
        days_open = t['days_open'] if t['days_open'] != -1 else '-'
        data_row = [
            t['ticket_id'], t['title'], t['status'], t['priority'],
            t['request_type'], labels_str, t['staff_assigned'], t['requester'],
            t['date_opened'], days_open, t['description'], t['resolution_notes']
        ]
        for col_idx, value in enumerate(data_row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

        # Apply priority color (column 4)
        priority_cell = ws.cell(row=row_idx, column=4)
        if t['priority'] in priority_fills:
            priority_cell.fill = priority_fills[t['priority']]
            priority_cell.font = priority_fonts.get(t['priority'], Font())

        # Apply status color (column 3)
        status_cell = ws.cell(row=row_idx, column=3)
        if t['status'] in status_fills:
            status_cell.fill = status_fills[t['status']]

    # Auto-filter
    if len(tickets) > 0:
        ws.auto_filter.ref = ws.dimensions

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 3, 50)

    # --- Sheet 2: Summary ---
    ws2 = wb.create_sheet("Summary")
    stats = calculate_stats(tickets)

    title_font = Font(bold=True, size=14, color='1a1a1a')
    section_font = Font(bold=True, size=12, color='333333')
    meta_font = Font(color='666666', size=10)

    ws2.cell(row=1, column=1, value="IT Service Tickets - Summary Report").font = title_font
    ws2.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = meta_font
    ws2.cell(row=3, column=1, value=f"Total Tickets: {stats['total']}").font = Font(bold=True, size=11)

    # Status breakdown
    row = 5
    ws2.cell(row=row, column=1, value="Status Breakdown").font = section_font
    ws2.cell(row=row, column=1).fill = PatternFill(start_color='f3f4f6', end_color='f3f4f6', fill_type='solid')
    ws2.cell(row=row, column=2).fill = PatternFill(start_color='f3f4f6', end_color='f3f4f6', fill_type='solid')
    row += 1
    for status, count in sorted(stats['statuses'].items(), key=lambda x: x[1], reverse=True):
        ws2.cell(row=row, column=1, value=status)
        ws2.cell(row=row, column=2, value=count)
        if status in status_fills:
            ws2.cell(row=row, column=1).fill = status_fills[status]
        row += 1

    # Priority breakdown
    row += 1
    ws2.cell(row=row, column=1, value="Priority Breakdown").font = section_font
    ws2.cell(row=row, column=1).fill = PatternFill(start_color='f3f4f6', end_color='f3f4f6', fill_type='solid')
    ws2.cell(row=row, column=2).fill = PatternFill(start_color='f3f4f6', end_color='f3f4f6', fill_type='solid')
    row += 1
    for priority, count in sorted(stats['priorities'].items(), key=lambda x: x[1], reverse=True):
        ws2.cell(row=row, column=1, value=priority)
        ws2.cell(row=row, column=2, value=count)
        if priority in priority_fills:
            ws2.cell(row=row, column=1).fill = priority_fills[priority]
        row += 1

    # Request types
    row += 1
    ws2.cell(row=row, column=1, value="Request Types").font = section_font
    ws2.cell(row=row, column=1).fill = PatternFill(start_color='f3f4f6', end_color='f3f4f6', fill_type='solid')
    ws2.cell(row=row, column=2).fill = PatternFill(start_color='f3f4f6', end_color='f3f4f6', fill_type='solid')
    row += 1
    for rt, count in sorted(stats['request_types'].items(), key=lambda x: x[1], reverse=True):
        ws2.cell(row=row, column=1, value=rt)
        ws2.cell(row=row, column=2, value=count)
        row += 1

    # Staff workload
    row += 1
    ws2.cell(row=row, column=1, value="Staff Workload").font = section_font
    ws2.cell(row=row, column=1).fill = PatternFill(start_color='f3f4f6', end_color='f3f4f6', fill_type='solid')
    ws2.cell(row=row, column=2).fill = PatternFill(start_color='f3f4f6', end_color='f3f4f6', fill_type='solid')
    ws2.cell(row=row, column=3).fill = PatternFill(start_color='f3f4f6', end_color='f3f4f6', fill_type='solid')
    row += 1
    ws2.cell(row=row, column=1, value="Staff Member").font = Font(bold=True)
    ws2.cell(row=row, column=2, value="Assigned").font = Font(bold=True)
    ws2.cell(row=row, column=3, value="Open").font = Font(bold=True)
    row += 1
    for staff, data in sorted(stats['staff_workload'].items()):
        ws2.cell(row=row, column=1, value=staff)
        ws2.cell(row=row, column=2, value=data['assigned'])
        ws2.cell(row=row, column=3, value=data['open'])
        row += 1

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 12

    return wb

# --- ROUTES ---
@app.route('/login')
def login_page():
    if 'user' in session:
        return redirect('/')
    return render_template('login.html')

@app.route('/api/login', methods=['POST'])
def api_login():
    data = request.json
    username = data.get('username', '').strip()
    password = data.get('password', '')

    user = get_user(username)
    if user and verify_password(password, user['password_hash']):
        session['user'] = {
            'username': user['username'],
            'role': user['role'],
            'display_name': user['display_name']
        }
        return jsonify({'status': 'success', 'user': session['user']})
    return jsonify({'error': 'Invalid credentials'}), 401

@app.route('/api/logout', methods=['POST'])
def api_logout():
    session.pop('user', None)
    return jsonify({'status': 'success'})

@app.route('/api/me')
def api_me():
    if 'user' in session:
        return jsonify(session['user'])
    return jsonify({'error': 'Not authenticated'}), 401

@app.route('/api/users', methods=['POST'])
@admin_required
def api_create_user():
    data = request.json
    username = data.get('username', '').strip()
    password = data.get('password', '')
    role = data.get('role', 'viewer')
    display_name = data.get('display_name', username)

    if not username or not password:
        return jsonify({'error': 'Username and password required'}), 400

    if get_user(username):
        return jsonify({'error': 'Username already exists'}), 400

    users = load_users()
    users.append({
        'username': username,
        'password_hash': hash_password(password),
        'role': role,
        'display_name': display_name
    })
    save_users(users)
    return jsonify({'status': 'success'}), 201

@app.route('/api/users', methods=['GET'])
@admin_required
def api_list_users():
    users = load_users()
    return jsonify([{
        'username': u['username'],
        'role': u['role'],
        'display_name': u['display_name']
    } for u in users])

@app.route('/api/users/<username>', methods=['PUT'])
@admin_required
def api_update_user(username):
    data = request.json
    users = load_users()

    for user in users:
        if user['username'] == username:
            if 'role' in data:
                user['role'] = data['role']
            if 'display_name' in data:
                user['display_name'] = data['display_name']
            if 'password' in data and data['password']:
                user['password_hash'] = hash_password(data['password'])
            save_users(users)
            return jsonify({'status': 'success'})
    return jsonify({'error': 'User not found'}), 404

@app.route('/api/users/<username>', methods=['DELETE'])
@admin_required
def api_delete_user(username):
    if username == 'admin':
        return jsonify({'error': 'Cannot delete admin user'}), 400

    users = load_users()
    users = [u for u in users if u['username'] != username]
    save_users(users)
    return jsonify({'status': 'success'})

@app.route('/api/change-password', methods=['POST'])
@login_required
def api_change_password():
    data = request.json
    current_password = data.get('current_password', '')
    new_password = data.get('new_password', '')

    if not new_password:
        return jsonify({'error': 'New password required'}), 400

    users = load_users()
    username = session['user']['username']

    for user in users:
        if user['username'] == username:
            if not verify_password(current_password, user['password_hash']):
                return jsonify({'error': 'Current password incorrect'}), 401
            user['password_hash'] = hash_password(new_password)
            save_users(users)
            return jsonify({'status': 'success'})
    return jsonify({'error': 'User not found'}), 404

# --- CATEGORY CRUD ---
@app.route('/api/categories')
@login_required
def api_categories():
    conn = get_db_connection()
    if not conn:
        return jsonify([])
    try:
        cur = conn.cursor()
        cur.execute('SELECT id, name, color, icon, is_custom, is_active FROM categories WHERE is_active = TRUE ORDER BY sort_order, name')
        categories = [dict(c) for c in cur.fetchall()]
        cur.close()
        conn.close()
        return jsonify(categories)
    except Exception as e:
        print(f"[API] Error fetching categories: {e}")
        return jsonify([])

@app.route('/api/categories', methods=['POST'])
@admin_required
def api_create_category():
    data = request.json
    name = data.get('name', '').strip()
    color = data.get('color', '#6366f1')
    icon = data.get('icon', 'fa-tag')
    if not name:
        return jsonify({'error': 'Name required'}), 400
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database error'}), 500
    try:
        cur = conn.cursor()
        cur.execute('''INSERT INTO categories (name, color, icon, is_custom)
                       VALUES (%s, %s, %s, TRUE) RETURNING id, name, color, icon''',
                    (name, color, icon))
        cat = dict(cur.fetchone())
        conn.commit()
        cur.close()
        conn.close()
        return jsonify(cat), 201
    except Exception as e:
        conn.rollback()
        conn.close()
        if 'unique' in str(e).lower() or 'duplicate' in str(e).lower():
            return jsonify({'error': 'Category already exists'}), 409
        return jsonify({'error': str(e)}), 500

@app.route('/api/categories/<int:cat_id>', methods=['PUT'])
@admin_required
def api_update_category(cat_id):
    data = request.json
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database error'}), 500
    try:
        cur = conn.cursor()
        updates, values = [], []
        for field in ['name', 'color', 'icon', 'is_active']:
            if field in data:
                updates.append(f"{field} = %s")
                values.append(data[field])
        if not updates:
            cur.close()
            conn.close()
            return jsonify({'status': 'no changes'}), 200
        values.append(cat_id)
        cur.execute(f"UPDATE categories SET {', '.join(updates)} WHERE id = %s", values)
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({'status': 'success'})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/categories/<int:cat_id>', methods=['DELETE'])
@admin_required
def api_delete_category(cat_id):
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database error'}), 500
    try:
        cur = conn.cursor()
        cur.execute('UPDATE categories SET is_active = FALSE WHERE id = %s', (cat_id,))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({'status': 'success'})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 500

# --- LABEL CRUD ---
@app.route('/api/labels')
@login_required
def api_labels():
    conn = get_db_connection()
    if not conn:
        return jsonify([])
    try:
        cur = conn.cursor()
        cur.execute('SELECT id, name, color FROM labels ORDER BY name')
        labels = [dict(l) for l in cur.fetchall()]
        cur.close()
        conn.close()
        return jsonify(labels)
    except Exception as e:
        print(f"[API] Error fetching labels: {e}")
        return jsonify([])

@app.route('/api/labels', methods=['POST'])
@admin_required
def api_create_label():
    data = request.json
    name = data.get('name', '').strip().upper()
    color = data.get('color', '#3b82f6')
    if not name:
        return jsonify({'error': 'Name required'}), 400
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database error'}), 500
    try:
        cur = conn.cursor()
        cur.execute('INSERT INTO labels (name, color) VALUES (%s, %s) RETURNING id, name, color',
                    (name, color))
        label = dict(cur.fetchone())
        conn.commit()
        cur.close()
        conn.close()
        return jsonify(label), 201
    except Exception as e:
        conn.rollback()
        conn.close()
        if 'unique' in str(e).lower() or 'duplicate' in str(e).lower():
            return jsonify({'error': 'Label already exists'}), 409
        return jsonify({'error': str(e)}), 500

@app.route('/api/labels/<int:label_id>', methods=['DELETE'])
@admin_required
def api_delete_label(label_id):
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database error'}), 500
    try:
        cur = conn.cursor()
        cur.execute('DELETE FROM labels WHERE id = %s', (label_id,))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({'status': 'success'})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 500

# --- TICKET LABELS ---
@app.route('/api/tickets/<ticket_id>/labels', methods=['PUT'])
@admin_required
def api_set_ticket_labels(ticket_id):
    data = request.json
    label_ids = data.get('label_ids', [])
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database error'}), 500
    try:
        cur = conn.cursor()
        cur.execute('DELETE FROM ticket_labels WHERE ticket_id = %s', (ticket_id,))
        for lid in label_ids:
            cur.execute('INSERT INTO ticket_labels (ticket_id, label_id) VALUES (%s, %s) ON CONFLICT DO NOTHING',
                        (ticket_id, int(lid)))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({'status': 'success'})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 500

# --- KANBAN BOARD ---
@app.route('/api/kanban')
@login_required
def api_kanban():
    tickets = read_tickets_from_db()
    columns = {
        'TO DO': [],
        'IN PROGRESS': [],
        'IN REVIEW': [],
        'DONE': []
    }
    status_map = {
        'Open': 'TO DO',
        'In Progress': 'IN PROGRESS',
        'Waiting for Info': 'IN REVIEW',
        'Resolved': 'DONE',
        'Closed': 'DONE'
    }
    for t in tickets:
        if t['days_open'] == -1:
            t['days_open'] = '-'
        col = status_map.get(t['status'], 'TO DO')
        columns[col].append(t)
    return jsonify(columns)

@app.route('/api/tickets/<ticket_id>/move', methods=['PUT'])
@admin_required
def api_move_ticket(ticket_id):
    data = request.json
    target_column = data.get('column', '')
    column_to_status = {
        'TO DO': 'Open',
        'IN PROGRESS': 'In Progress',
        'IN REVIEW': 'Waiting for Info',
        'DONE': 'Resolved'
    }
    new_status = column_to_status.get(target_column)
    if not new_status:
        return jsonify({'error': 'Invalid column'}), 400
    if update_ticket_in_db(ticket_id, {'status': new_status}):
        return jsonify({'status': 'success', 'new_status': new_status})
    return jsonify({'error': 'Ticket not found'}), 404

# --- TICKET ATTACHMENTS ---
@app.route('/api/tickets/<ticket_id>/attachments', methods=['POST'])
@admin_required
def api_upload_attachment(ticket_id):
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    # Validate file type
    allowed_types = {'image/png', 'image/jpeg', 'image/gif', 'image/webp', 'image/svg+xml',
                     'application/pdf', 'text/plain'}
    if file.content_type not in allowed_types:
        return jsonify({'error': 'File type not allowed. Accepted: images, PDF, text'}), 400
    # Limit file size (5MB)
    file_data = file.read()
    if len(file_data) > 5 * 1024 * 1024:
        return jsonify({'error': 'File too large (max 5MB)'}), 400
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database error'}), 500
    try:
        cur = conn.cursor()
        import uuid
        filename = f"{uuid.uuid4().hex}_{file.filename}"
        cur.execute('''INSERT INTO ticket_attachments (ticket_id, filename, original_name, mime_type, file_data, file_size)
                       VALUES (%s, %s, %s, %s, %s, %s) RETURNING id, filename, original_name, mime_type, file_size''',
                    (ticket_id, filename, file.filename, file.content_type, psycopg2.Binary(file_data), len(file_data)))
        attachment = dict(cur.fetchone())
        conn.commit()
        cur.close()
        conn.close()
        return jsonify(attachment), 201
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/tickets/<ticket_id>/attachments')
@login_required
def api_get_attachments(ticket_id):
    conn = get_db_connection()
    if not conn:
        return jsonify([])
    try:
        cur = conn.cursor()
        cur.execute('''SELECT id, filename, original_name, mime_type, file_size, created_at
                       FROM ticket_attachments WHERE ticket_id = %s ORDER BY created_at''', (ticket_id,))
        attachments = []
        for row in cur.fetchall():
            att = dict(row)
            att['created_at'] = att['created_at'].isoformat() if att['created_at'] else ''
            attachments.append(att)
        cur.close()
        conn.close()
        return jsonify(attachments)
    except Exception as e:
        print(f"[API] Error fetching attachments: {e}")
        return jsonify([])

@app.route('/api/attachments/<int:attachment_id>')
@login_required
def api_serve_attachment(attachment_id):
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database error'}), 500
    try:
        cur = conn.cursor()
        cur.execute('SELECT filename, original_name, mime_type, file_data FROM ticket_attachments WHERE id = %s',
                    (attachment_id,))
        row = cur.fetchone()
        cur.close()
        conn.close()
        if not row:
            return jsonify({'error': 'Attachment not found'}), 404
        return send_file(
            BytesIO(bytes(row['file_data'])),
            mimetype=row['mime_type'],
            download_name=row['original_name']
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/attachments/<int:attachment_id>', methods=['DELETE'])
@admin_required
def api_delete_attachment(attachment_id):
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database error'}), 500
    try:
        cur = conn.cursor()
        cur.execute('DELETE FROM ticket_attachments WHERE id = %s', (attachment_id,))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({'status': 'success'})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/')
def index():
    if 'user' not in session:
        return redirect('/login')
    return render_template('dashboard.html')

@app.route('/api/stats')
def api_stats():
    tickets = read_tickets_from_db()
    stats = calculate_stats(tickets)
    return jsonify(stats)

@app.route('/api/options')
def api_options():
    return jsonify(get_dropdown_options())

@app.route('/api/tickets')
def api_tickets():
    tickets = read_tickets_from_db()
    # Convert -1 sentinel to '-' for closed/resolved tickets
    for t in tickets:
        if t['days_open'] == -1:
            t['days_open'] = '-'
    return jsonify(tickets)

@app.route('/api/tickets', methods=['POST'])
def create_ticket():
    try:
        data = request.json
        ticket_id = create_ticket_in_db(data)
        if ticket_id:
            return jsonify({'status': 'success', 'ticket_id': ticket_id}), 201
        return jsonify({'error': 'Failed to save'}), 500
    except Exception as e:
        print(f"[API] Error creating ticket: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/tickets/<ticket_id>', methods=['PUT'])
def update_ticket(ticket_id):
    try:
        data = request.json
        if update_ticket_in_db(ticket_id, data):
            return jsonify({'status': 'success'})
        return jsonify({'error': 'Ticket not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/tickets/<ticket_id>', methods=['DELETE'])
def delete_ticket(ticket_id):
    try:
        if delete_ticket_from_db(ticket_id):
            return jsonify({'status': 'success'})
        return jsonify({'error': 'Ticket not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/refresh', methods=['POST'])
def api_refresh():
    return jsonify({'status': 'success', 'message': 'Data refreshed from database'})

@app.route('/api/export')
def api_export():
    """Export tickets to Excel file."""
    try:
        wb = generate_excel_from_db()
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"tickets_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"[API] Error exporting: {e}")
        return jsonify({'error': str(e)}), 500

# Startup diagnostics
def print_diagnostics():
    """Print startup diagnostics."""
    print("\n" + "="*50)
    print(" IT OPSCENTER - STARTUP DIAGNOSTICS")
    print("="*50)

    # Database check
    print("\n[DATABASE]")
    if DATABASE_URL:
        print(f"  DATABASE_URL: {'*' * 20}...{DATABASE_URL[-20:]}")
        print("  Initializing database...")
        if init_database():
            print("  Database: OK")
        else:
            print("  Database: FAILED")
    else:
        print("  DATABASE_URL: NOT SET - using Excel fallback")

    # Auth check
    print("\n[AUTHENTICATION]")
    for path in USERS_FILE_LOCATIONS:
        exists = os.path.exists(path)
        status = "FOUND" if exists else "not found"
        print(f"  {path}: {status}")

    users_file = get_users_file()
    print(f"\n  Active users file: {users_file}")

    users = load_users()
    if users:
        print(f"  Users loaded: {len(users)}")
        for u in users:
            print(f"    - {u.get('username')} ({u.get('role')})")
    else:
        print("  WARNING: No users loaded! Login will fail.")

    print("="*50 + "\n")

# Run diagnostics on startup
print_diagnostics()

if __name__ == '__main__':
    print("\n" + "="*50)
    print(" IT OPERATIONS CENTER - DASHBOARD")
    print(" Open browser to: http://127.0.0.1:5000")
    print("="*50 + "\n")
    app.run(debug=True, port=5000)
